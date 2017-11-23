from openpyxl import load_workbook

class ExcelReader(object):
    '''iterator that reads an excel file line by line'''

    def __init__(self, reader, sheet_name, col_names=None, first_data_row = 2):
        self.col_names = col_names
        self.reader = reader
        # self.workbook = load_workbook('sources/Quickbooks_FactFinance.xlsx', data_only=True)
        self.workbook = load_workbook(filename = reader.getFile(), data_only=True)
        #Specifying the sheet name we want to read.
        self.sheet = self.workbook.get_sheet_by_name(sheet_name)
        self.max_row = self.sheet.max_row
        self.max_col = len(self.col_names)
        self.first_data_row = first_data_row

    def rows(self):
        for row in range(self.first_data_row, self.max_row + 1):
            result = self.readRow(row)
            for val in result.values():
                if val != None:
                    yield result
                    break

    def readRow(self, row):
        result = {}
        for col in range(1, self.max_col + 1):
            value = self.sheet.cell(row=row, column=col).value
            if value == None or value == "NULL" or value == "":
                value = None
            else:
                value = str(value)
            result[self.col_names[col-1]] = value
        return result

    def columns(self):
        return self.readRow(1)

    def close(self):
        self.workbook.close()